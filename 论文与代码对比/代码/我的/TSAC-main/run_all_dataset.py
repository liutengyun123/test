import cv2
import numpy as np
from openpyxl import load_workbook
import glob
import os
import run

#数据集存放的位置
folder_path = r'E:\paper_testcode\match\dataset\test_vgg'
excel1 = r'E:\paper_testcode\match\dataset\result_vgg.xlsx'

# folder_path = r'E:\paper_testcode\match\dataset\test'
# excel1 = r'E:\paper_testcode\match\dataset\result.xlsx'

print("开始")

folders = glob.glob(folder_path+'/*')
wb=load_workbook(excel1)
sheet=wb['消融实验2']        #表格中的sheet
i=2    
recall_sum = 0
precison_sum = 0
F_sum = 0
count = 0
for item in folders:
    cls = os.path.basename(item)
    print(cls)

    for filename in os.listdir(item):
        if filename == 'H_1_3' or filename == 'H1to2p':
            H = []  
            file = os.path.join(item, filename)
            with open(file, 'rb') as f: 
                lines = f.readlines() 
                
                for line in lines:
                    l = str(line).split('b')[1].split("'")[1].split("\\")[0].split()
                    h = []
                    for element in l:
                        element = float(element)
                        h.append(element)
                    H.append(h)
            if len(H) >3:
                H = H[:3]
            print(H)
            # img1 = os.path.join(item,'1.ppm')
            # img2 = os.path.join(item,'3.ppm')
            #如果是vgg数据集，跑这段代码
            img1 = os.path.join(item,'img1.ppm')
            img2 = os.path.join(item,'img2.ppm')
            if os.path.isfile(img1) == False:  
                img1 = os.path.join(item,'img1.pgm')
                img2 = os.path.join(item,'img2.pgm')
            #print(img1,img2)
            recall,precision,F=run.run(img1,img2,H)
            recall_sum = recall_sum + recall
            precison_sum = precison_sum + precision
            F_sum = F_sum + F
            count = count+1
            sheet.cell(row=i, column=1, value=cls)
            sheet.cell(row=i, column=2, value=recall)   
            sheet.cell(row=i, column=3, value=precision)  
            sheet.cell(row=i, column=4, value=F)
            i = i+1 
            print("......................................................................") 
sheet.cell(row=i, column=1, value="平均")
sheet.cell(row=i, column=2, value=recall_sum/count)   
sheet.cell(row=i, column=3, value=precison_sum/count)  
sheet.cell(row=i, column=4, value=F_sum/count)
wb.save(excel1)
print("结束")
    
   